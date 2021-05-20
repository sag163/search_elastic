import fitz
import openpyxl
import xlrd
import textract
from striprtf.striprtf import rtf_to_text


# def parser_pdf(link: str) -> str: # пока не нужен, можно использовать parser_doc
#     ''' Функция извлечения текста из pdf файла'''
#     doc = fitz.open(link)
#     text_list = []
#     count = 0
#     for current_page in range(len(doc)):
#         if count < 100_000:
#             page = doc.loadPage(current_page)
#             page_text = page.getText("text")
#             text_list.append(page_text)
#             count += len(page_text)
#         else:
#             break
#     text = ' '.join(text_list)
#     return text


def parser_default(link: str) -> str:
    '''Функция извлечения текста из файлов не требующих индивидуального
    подхода'''
    with open(link, encoding='utf-8') as f:
        text = f.read()
        if len(text) > 100_000:
            description = text[:100_000].rstrip()
        else:
            description = text.rstrip()
    return description


def parser_xlxs(link: str) -> str:
    '''Функция извлечения текста из файлов exel формата xlxs'''
    workbook = openpyxl.load_workbook(link)
    sheets = workbook.sheetnames
    text_list = []
    count = 0
    for name in sheets:
        sheet = workbook.get_sheet_by_name(name)
        rows = sheet.max_row
        cols = sheet.max_column
        for i in range(1, rows + 1):
            if count < 100_000:
                for j in range(1, cols + 1):
                    cell = sheet.cell(row=i, column=j)
                    cell_text = str(cell.value)
                    text_list.append(cell_text)
                    count += len(cell_text)
            else:
                break
    text = ' '.join(text_list)
    return text


def parser_xlx(link: str) -> str:
    '''Функция извлечения текста из файлов exel формата xlx'''
    workbook = xlrd.open_workbook(link)
    sheets = workbook.sheet_names()
    text_list = []
    count = 0
    for name in sheets:
        sheet = workbook.sheet_by_name(name)
        rows = sheet.nrows
        cols = sheet.ncols
        for i in range(rows):
            if count < 100_000:
                for j in range(cols):
                    cell = sheet.cell_value(rowx=i, colx=j)
                    cell_text = str(cell)
                    text_list.append(cell_text)
                    count += len(cell_text)
            else:
                break
    text = ' '.join(text_list)
    return text


def parser_doc(link: str) -> str:
    '''Функция извлечения текста из файлов .doc и .docx'''
    text = textract.process(link)
    print(text)
    text = text.decode("utf-8")
    print(type(text))
    if len(text) > 100_000:
        description = text[:100_000].rstrip()
    else:
        description = text.rstrip()
    return description


def parser_rtf(link: str) -> str:
    '''Функция извлечения текста из файлов .rtf'''
    with open(link, encoding='utf-8') as f:
        text = f.read()
        if len(text) > 100_000:
            description = text[:100_000].rstrip()
        else:
            description = text.rstrip()
    text = rtf_to_text(description)
    return text

